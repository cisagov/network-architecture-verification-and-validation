#!/usr/bin/env python3

# Copyright 2023 Battelle Energy Alliance, LLC

import os
import itertools
from collections import Counter
import socket
from copy import copy
import json
import pkg_resources
import pickle
import string

import openpyxl
import openpyxl.styles
from openpyxl.worksheet.table import Table
import netaddr
from tqdm import tqdm

from navv import data_types
from navv.utilities import timeit
from navv.message_handler import warning_msg


DATA_PKL_FILE = pkg_resources.resource_filename(__name__, "data/data.pkl")
COL_NAMES = [
    "Count",
    "Src_IP",
    "Src_Desc",
    "Dest_IP",
    "Dest_Desc",
    "Port",
    "Service",
    "Proto",
    "Conn_State",
    "Notes",
]
HEADER_STYLE = openpyxl.styles.NamedStyle(
    name="header_style",
    font=openpyxl.styles.Font(name="Calibri", size=11, bold=True),
    fill=openpyxl.styles.PatternFill("solid", fgColor="4286F4"),
)
IPV6_CELL_COLOR = (
    openpyxl.styles.PatternFill("solid", fgColor="FFFFFF"),
    openpyxl.styles.Font(name="Calibri", size=11, color="ff0000"),
)
EXTERNAL_NETWORK_CELL_COLOR = (
    openpyxl.styles.PatternFill("solid", fgColor="030303"),
    openpyxl.styles.Font(name="Calibri", size=11, color="ffff00"),
)
INTERNAL_NETWORK_CELL_COLOR = (
    openpyxl.styles.PatternFill("solid", fgColor="ffff00"),
    openpyxl.styles.Font(name="Calibri", size=11, color="000000"),
)
ICMP_CELL_COLOR = (
    openpyxl.styles.PatternFill("solid", fgColor="ff33cc"),
    openpyxl.styles.Font(name="Calibri", size=11, color="000000"),
)
UNKNOWN_EXTERNAL_CELL_COLOR = (
    openpyxl.styles.PatternFill("solid", fgColor="ffffff"),
    openpyxl.styles.Font(name="Calibri", size=11, color="000000"),
)
ALREADY_UNRESOLVED = list()


def get_workbook(file_name):
    """Create the blank Inventory and Segment sheets for data input into the tool"""
    if os.path.isfile(file_name):
        wb = openpyxl.load_workbook(file_name)
    else:
        wb = openpyxl.Workbook()
        inv_sheet = wb.active
        inv_sheet.title = "Inventory Input"
        seg_sheet = wb.create_sheet("Segments")

        inv_sheet.cell(row=1, column=1, value="IP").style = HEADER_STYLE
        inv_sheet.cell(row=1, column=2, value="Name").style = HEADER_STYLE

        seg_sheet.cell(row=1, column=1, value="Name").style = HEADER_STYLE
        seg_sheet.cell(row=1, column=2, value="Description").style = HEADER_STYLE
        seg_sheet.cell(row=1, column=3, value="CIDR").style = HEADER_STYLE
    return wb


@timeit
def get_inventory_data(ws, **kwargs):
    inventory = dict()
    for row in itertools.islice(ws.iter_rows(), 1, None):
        if not row[0].value or not row[1].value:
            continue
        inventory[row[0].value] = data_types.InventoryItem(
            ip=row[0].value,
            name=row[1].value,
            color=(copy(row[0].fill), copy(row[0].font)),
            mac_address="",
            vendor=""
        )
    return inventory


@timeit
def get_segments_data(ws):
    segments = []
    network_ip = ""
    for row in itertools.islice(ws.iter_rows(), 1, None):
        if not row[2].value:
            continue
        network_ip = row[2].value
        network_ips = [str(ip) for ip in netaddr.IPNetwork(network_ip)]
        for ip in network_ips:
            segments.append(
                data_types.Segment(
                    name=row[0].value,
                    description=row[1].value,
                    network=ip,
                    color=[copy(row[0].fill), copy(row[0].font)],
                )
            )
    return segments


def get_package_data():
    """Load services and conn_states data into memory"""
    with open(DATA_PKL_FILE, "rb") as f:
        services, conn_states = pickle.load(f)
    return services, conn_states


@timeit
def create_analysis_array(sort_input, **kwargs):
    arr = []
    # sort by count and source IP
    counted = sorted(
        list(
            str(count) + "\t" + item
            for item, count in sorted(Counter(sort_input).items(), key=lambda x: x[0])
        ),
        key=lambda x: int(x.split("\t")[0]),
        reverse=True,
    )
    for row in counted:
        cells = row.split("\t")
        arr.append(
            data_types.AnalysisRowItem(
                count=cells[0],
                src_ip=cells[1],
                dest_ip=cells[2],
                port=cells[3],
                proto=cells[4],
                conn=cells[5],
            )
        )

    return arr


@timeit
def perform_analysis(
    wb,
    rows,
    services,
    conn_states,
    inventory,
    segments,
    dns_data,
    json_path,
    ext_IPs,
    unk_int_IPs,
    **kwargs,
):
    sheet = make_sheet(wb, "Analysis", idx=0)
    sheet.append(
        [
            "Count",
            "Src_IP",
            "Src_Desc",
            "Dest_IP",
            "Dest_Desc",
            "Port",
            "Service",
            "Proto",
            "Conn_State",
            "Notes",
        ]
    )
    warning_msg("this may take awhile...")
    for row_index, row in enumerate(tqdm(rows), start=2):
        row.src_desc = handle_ip(
            row.src_ip, dns_data, inventory, segments, ext_IPs, unk_int_IPs
        )
        row.dest_desc = handle_ip(
            row.dest_ip, dns_data, inventory, segments, ext_IPs, unk_int_IPs
        )
        handle_service(row, services)
        row.conn = (row.conn, conn_states[row.conn])
        write_row_to_sheet(row, row_index, sheet)
    tab = Table(displayName="AnalysisTable", ref=f"A1:J{len(rows)+1}")
    sheet.add_table(tab)
    # write lookup data to json file for future use
    with open(json_path, "w+") as fp:
        json.dump(dns_data, fp)


def write_row_to_sheet(row, row_index, sheet):
    sheet.cell(row=row_index, column=1, value=int(row.count))

    src_IP = sheet.cell(row=row_index, column=2, value=row.src_ip)
    src_IP.fill = row.src_desc[1][0]
    src_IP.font = row.src_desc[1][1]

    src_Desc = sheet.cell(row=row_index, column=3, value=row.src_desc[0])
    src_Desc.fill = row.src_desc[1][0]
    src_Desc.font = row.src_desc[1][1]

    dest_IP = sheet.cell(row=row_index, column=4, value=row.dest_ip)
    dest_IP.fill = row.dest_desc[1][0]
    dest_IP.font = row.dest_desc[1][1]

    dest_Desc = sheet.cell(row=row_index, column=5, value=row.dest_desc[0])
    dest_Desc.fill = row.dest_desc[1][0]
    dest_Desc.font = row.dest_desc[1][1]

    sheet.cell(row=row_index, column=6, value=int(row.port))

    service = sheet.cell(row=row_index, column=7, value=row.service[0])
    service.fill = row.service[1][0]
    service.font = row.service[1][1]

    sheet.cell(row=row_index, column=8, value=row.proto)

    conn_State = sheet.cell(row=row_index, column=9, value=row.conn[0])
    conn_State.fill = row.conn[1][0]
    conn_State.font = row.conn[1][1]

    # placeholder for notes cell
    sheet.cell(row=row_index, column=10, value="")


def handle_service(row, services):
    # { port: { proto: (name, (fill, font)} }
    if row.port in services and row.proto in services[row.port]:
        row.service = services[row.port][row.proto]
    else:
        if row.proto == "icmp":
            if netaddr.valid_ipv4(row.src_ip):
                row.proto = "ICMPv4"
                service_dict = data_types.icmp4_types
            else:
                row.proto = "ICMPv6"
                service_dict = data_types.icmp6_types
            if row.port in service_dict:
                row.service = (service_dict[row.port], ICMP_CELL_COLOR)
            else:
                row.service = ("unknown icmp", ICMP_CELL_COLOR)
        else:
            row.service = ("unknown service", UNKNOWN_EXTERNAL_CELL_COLOR)


def handle_ip(ip_to_check, dns_data, inventory, segments, ext_IPs, unk_int_IPs):
    """Function take IP Address and uses collected dns_data, inventory, and segment information to give IP Addresses in analysis context.

    Priority flow:
        * DHCP Broadcasting
        * Multicast
        * Within Segments identified
            * Resolution by DNS, then Inventory, and then Unknown
            * Appends name if External IP
        * Private Network
            * Resolution by DNS, Inventory, then Unknown
        * External (Public IP space) or Internet
            * Resolution by DNS, Unknown

    This will capture the name description and the color coding identified within the worksheet.
    """
    segment_ips = [segment.network for segment in segments]
    desc_to_change = ("Not Triggered IP", IPV6_CELL_COLOR)
    if ip_to_check == str("0.0.0.0"):
        desc_to_change = (
            "Unassigned IPv4",
            IPV6_CELL_COLOR,
        )
    elif ip_to_check == str("255.255.255.255"):
        desc_to_change = (
            "IPv4 All Subnet Broadcast",
            IPV6_CELL_COLOR,
        )
    elif (
        netaddr.valid_ipv6(ip_to_check) or netaddr.IPAddress(ip_to_check).is_multicast()
    ):
        desc_to_change = (
            f"{'IPV6' if netaddr.valid_ipv6(ip_to_check) else 'IPV4'}{'_Multicast' if netaddr.IPAddress(ip_to_check).is_multicast() else ''}",
            IPV6_CELL_COLOR,
        )
    elif ip_to_check in segment_ips:
        for x in range(0, len(segments[:-1])):
            if segments[x].network == ip_to_check:
                if ip_to_check in dns_data:
                    resolution = dns_data[ip_to_check].name
                elif ip_to_check in inventory:
                    resolution = inventory[ip_to_check].name
                else:
                    resolution = f"Unknown device in {segments[x].name} network"
                    unk_int_IPs.add(ip_to_check)
                if not netaddr.IPAddress(ip_to_check).is_private():
                    resolution = resolution + " {Non-Priv IP}"
                desc_to_change = (
                    resolution,
                    segments[x].color,
                    )
    elif netaddr.IPAddress(ip_to_check).is_private():
        if ip_to_check in dns_data:
            desc_to_change = (dns_data[ip_to_check], INTERNAL_NETWORK_CELL_COLOR)
        elif ip_to_check in inventory:
            desc_to_change = (inventory[ip_to_check].name, INTERNAL_NETWORK_CELL_COLOR)
        else:
            desc_to_change = ("Unknown Internal address", INTERNAL_NETWORK_CELL_COLOR)
            unk_int_IPs.add(ip_to_check)
    else:
        ext_IPs.add(ip_to_check)
        if ip_to_check in dns_data:
            resolution = dns_data[ip_to_check]
        elif ip_to_check in inventory:
            resolution = inventory[ip_to_check].name + " {Non-Priv IP}"
        else:
            try:
                resolution = socket.gethostbyaddr(ip_to_check)[0]
            except socket.herror:
                ALREADY_UNRESOLVED.append(ip_to_check)
            finally:
                resolution = "Unresolved external address"
        desc_to_change = (resolution, EXTERNAL_NETWORK_CELL_COLOR)
    return desc_to_change


def write_conn_states_sheet(conn_states, wb):
    new_ws = make_sheet(wb, "Conn States", idx=8)
    new_ws.append(["State", "Description"])
    for index, conn_state in enumerate(conn_states, start=2):
        # State column
        state_cell = new_ws[f"A{index}"]
        state_cell.value = conn_state
        state_cell.fill = conn_states[conn_state][0]
        state_cell.font = conn_states[conn_state][1]

        # Description column
        desc_cell = new_ws[f"B{index}"]
        desc_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        desc_cell.value = conn_states[conn_state][2]
        desc_cell.fill = conn_states[conn_state][0]
        desc_cell.font = conn_states[conn_state][1]
    auto_adjust_width(new_ws, 100)


def write_inventory_report_sheet(inventory_df, wb):
    """Get Mac Addresses with their associated IP addresses and manufacturer."""
    ir_sheet = make_sheet(wb, "Inventory Report", idx=4)
    ir_sheet.append(["MAC", "Vendor", "Hostname", "IPv4", "IPv6", "Port and Proto"])

    inventory_data = inventory_df.to_dict(orient="records")
    for index, row in enumerate(inventory_data, start=2):
        # Mac column
        ir_sheet[f"A{index}"].value = row["mac"]

        # Vendor column
        ir_sheet[f"B{index}"].value = row["vendor"]

        # Hostname column
        hostname_column = ir_sheet[f"C{index}"]
        hostname_column.alignment = openpyxl.styles.Alignment(wrap_text=True)

        hostname = ""
        if row["hostname"]:
            hostname = ", ".join(each for each in row["hostname"] if each)
        hostname_column.value = hostname

        # IPv4 Address column
        ipv4_column = ir_sheet[f"D{index}"]
        ipv4_column.alignment = openpyxl.styles.Alignment(wrap_text=True)

        ipv4 = ""
        if row["ipv4"]:
            ipv4 = ", ".join(each for each in row["ipv4"] if each)
        ipv4_column.value = ipv4

        # IPv6 Address column
        ipv6_column = ir_sheet[f"E{index}"]
        ipv6_column.alignment = openpyxl.styles.Alignment(wrap_text=True)

        ipv6 = ""
        if row["ipv6"]:
            ipv6 = ", ".join(each for each in row["ipv6"] if each)
        ipv6_column.value = ipv6

        # Port and Protocol column
        pnp_column = ir_sheet[f"F{index}"]
        pnp_column.alignment = openpyxl.styles.Alignment(wrap_text=True)

        port_and_proto = ""
        if row["port_and_proto"]:
            port_and_proto = ", ".join(
                list(set(each for each in row["port_and_proto"] if each))[:10]
            )

        pnp_column.value = port_and_proto

        # Add styling to every other row
        if index % 2 == 0:
            for cell in ir_sheet[f"{index}:{index}"]:
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor="AAAAAA")
    auto_adjust_width(ir_sheet, 40)


def write_snmp_sheet(snmp_df, wb):
    """Write SNMP log data to excel sheet."""
    sheet = make_sheet(wb, "SNMP", idx=4)
    sheet.append(
        ["Src IPv4", "Src Port", "Dest IPv4", "Dest Port", "Version", "Community"]
    )

    for index, row in enumerate(snmp_df.to_dict(orient="records"), start=2):
        # Source IPv4 column
        sheet[f"A{index}"].value = row["src_ip"]

        # Source Port column
        sheet[f"B{index}"].value = row["src_port"]

        # Destination IPv4 column
        sheet[f"C{index}"].value = row["dst_ip"]

        # Destination Port column
        sheet[f"D{index}"].value = row["dst_port"]

        # Version column
        sheet[f"E{index}"].value = row["version"]

        # Community column
        sheet[f"F{index}"].value = row["community"]

        # Add styling to every other row
        if index % 2 == 0:
            for cell in sheet[f"{index}:{index}"]:
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor="AAAAAA")

    auto_adjust_width(sheet, 40)


def write_externals_sheet(IPs, wb):
    ext_sheet = make_sheet(wb, "Externals", idx=5)
    ext_sheet.append(["External IP"])
    for row_index, IP in enumerate(sorted(IPs), start=2):
        cell = ext_sheet[f"A{row_index}"]
        cell.value = IP
        if row_index % 2 == 0:
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="AAAAAA")
    auto_adjust_width(ext_sheet)


def write_unknown_internals_sheet(IPs, wb):
    int_sheet = make_sheet(wb, "Unknown Internals", idx=6)
    int_sheet.append(["Unknown Internal IP"])
    for row_index, IP in enumerate(sorted(IPs), start=2):
        cell = int_sheet[f"A{row_index}"]
        cell.value = IP
        if row_index % 2 == 0:
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="AAAAAA")
    auto_adjust_width(int_sheet)


def write_stats_sheet(wb, stats):
    stats_sheet = make_sheet(wb, "Stats", idx=7)
    stats_sheet.append(
        ["Length of Capture time"]
        + [column for column in stats if column != "Length of Capture time"]
    )
    stats_sheet["A2"] = stats.pop("Length of Capture time")
    for col_index, stat in enumerate(stats, 1):
        stats_sheet[f"{string.ascii_uppercase[col_index]}2"].value = stats[stat]
    auto_adjust_width(stats_sheet)


def make_sheet(wb, sheet_name, idx=None):
    """Create the sheet if it doesn't already exist otherwise remove it and recreate it"""
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    return wb.create_sheet(sheet_name, index=idx)


def auto_adjust_width(sheet, width=40):
    """Adjust the width of the columns to fit the data"""
    for col in sheet.columns:
        max_width = max(len(f"{c.value}") for c in col if c.value) + 2
        sheet.column_dimensions[col[0].column_letter].width = (
            width if width < max_width else max_width
        )
